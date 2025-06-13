# -*- coding: utf-8 -*-
"""
generate_master_datasheet.py

This script automates the creation of the Master Equipment datasheet file from 
'Datasheets.xlsm', which contains multiple equipment-specific sheets.

Key Features:
- Creates the Master eexcel sheet with a time stamp in the name of the sheet
- Extracts parameter name (Column C), units (Column E), and category (Column I) from each sheet.
- Groups parameters into five standardized categories:
  - SysCAD Inputs, Engineering Inputs, Lab/Pilot Inputs, Project Constant, Vendor Inputs
- Generates one sheet per equipment with:
  - Header rows (equipment name, unit count placeholder, bold column titles)
  - Parameters grouped under each category
  - Category labels merged vertically
  - Auto-fit column widths
  

Sheets without valid category data are skipped. Output is saved as 'Master_DataSheet_Generated_vi.xlsx'.

Requirements: pandas, openpyxl

Created on Thu May 29 17:15:36 2025

@author: AsfiyaKhanam
"""
def generate_master_datasheet(uploaded_file):

    import warnings
    warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')


    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict
    from datetime import datetime
    from io import BytesIO
    #from pathlib import Path

    # --- CONFIGURATION ---
    #source_file = "Datasheets.xlsm"
    # output_file = "Master_DataSheet_Generated_v3.xlsx"
    # --- TIMESTAMPED OUTPUT FILE ---
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    output_filename = f"Master_DataSheet_{timestamp}.xlsx"


    # Category normalization map
    category_mapping = {
        "SysCAD": "SysCAD Inputs",
        "Engineering Input": "Engineering Inputs",
        "Lab/Pilot Value": "Lab/Pilot Inputs",
        "Project Constant": "Project Constant",
        "Vendor Input": "Vendor Inputs"
     }

    # Desired category order
    ordered_categories = [
        "SysCAD Inputs",
        "Engineering Inputs",
        "Lab/Pilot Inputs",
        "Project Constant",
        "Vendor Inputs"
    ]

    # Style Setup
    thin_border = Border(
        left = Side(style='thin'),
        right= Side(style="thin"),
        top = Side(style = 'thin'),
        bottom= Side(style = 'thin')
    )   

    # Load the Excel file
    xls = pd.ExcelFile(uploaded_file)
    sheet_names = xls.sheet_names

    # Create a new Excel workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for sheet_name in sheet_names:
        df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None)
    
        # Skip sheets without valid category data (column I)
        if df.shape[1] < 9 or df.iloc[:, 8].dropna().empty:
            continue
    
        # Extract parameter name (Column C), units (Column E), category (Column I)
        records = []
        for _, row in df.iterrows():
            param = row[2] if pd.notna(row[2]) else None
            unit = row[4] if pd.notna(row[4]) else ""
            category_raw = row[8] if pd.notna(row[8]) else None

            if param and category_raw:
                category = category_mapping.get(category_raw.strip(), None)
                if category:
                    records.append((category, str(param).strip(), str(unit).strip()))

        if not records:
            continue

        # Sort and group by category
        grouped = {cat: [] for cat in ordered_categories}
        for category, param, unit in records:
            grouped[category].append((param, unit))

        # Create a new sheet in the output workbook
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit

        # Write header rows
        ws.append([sheet_name])
        ws.append(["Number of units ="])
        header_row = ["Parameter Category", "Input Parameters", "Units"]
        ws.append(header_row)
    
        # Bold row 3
        for cell in ws[3]:
            if cell.value is not None:
                cell.font= Font(bold=True)
        
        # Track current row for parameter writing
        current_row = 4
    
        # Track column widths dynamically
        column_widths = defaultdict(int)
        for col_idx, value in enumerate(header_row,start=1):
            column_widths[col_idx] = max(column_widths[col_idx],len(str(value)))
            #cell = ws.cell(row=3,column=col_idx,value=value)
            #cell.border = thin_border #Applying border to the header
        
        # Write data, grouped by category  
        for category in ordered_categories:
            param_list = grouped[category]
            if not param_list:
                continue
        
            start_row = current_row
            # Write category + first parameter in same row
            first_param, first_unit = param_list[0]
            ws.cell(row=current_row, column=1, value=category)
            ws.cell(row=current_row, column=2, value=first_param)
            ws.cell(row=current_row, column=3, value=first_unit)
        
            column_widths[1] = max(column_widths[1],len(str(category)))
            column_widths[2] = max(column_widths[2],len(str(first_param)))
            column_widths[3] = max(column_widths[3],len(str(first_unit)))
        
            current_row += 1

            # Write remaining parameters
            for param, unit in param_list[1:]:
                ws.cell(row=current_row, column=2, value=param)
                ws.cell(row=current_row, column=3, value=unit)
                column_widths[2] = max(column_widths[2], len(str(param)))
                column_widths[3] = max(column_widths[3], len(str(unit)))
                current_row += 1
            
            end_row = current_row - 1
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

        # --- Auto-fit column widths ---
        for col_idx, width in column_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width + 4

        # --- APPLY BORDERS TO ALL NON-EMPTY CELLS ---
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None or cell.coordinate in ws.merged_cells:
                    cell.border = thin_border

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    print(f"✅ Master file ready for download: {output_filename}")
    return output, output_filename
