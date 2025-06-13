from io import BytesIO
from openpyxl import load_workbook
import pandas as pd
def populate_syscad_inputs(master_file: BytesIO, streamtable_file: BytesIO) -> BytesIO:

    # Load workbooks
    master_wb = load_workbook(master_file)
    streamtable_wb = load_workbook(streamtable_file, data_only=True)

    # Identify matching and missing equipment sheets
    master_equipment_types = set(master_wb.sheetnames)
    stream_equipment_types = set(streamtable_wb.sheetnames)
    matched_types = master_equipment_types & stream_equipment_types
    missing_types = master_equipment_types - stream_equipment_types

    for equipment_type in matched_types:
        master_sheet = master_wb[equipment_type]
        stream_sheet = streamtable_wb[equipment_type]

        # Parameter mapping: master label -> streamtable tag
        parameter_mapping = {
            "Operating Temperature": "Prod.T (C)",
            "Operating Density": "Prod.Rho (kg/m^3)",
            "Design Density": "Prod.Rho (kg/m^3)",
            "Operating Slurry pH": "Prod.pH",
            "Design Slurry pH": "Prod.pH",
            "Flow Rate to/from Vessel": "Prod.Qm (kg/h)"
        }

        # Get unit tags from streamtable (row 1, col D onward)
        stream_unit_tags = [cell.value for cell in stream_sheet[1][3:] if cell.value]

        # Populate unit tags into master sheet row 3 (D3 onward)
        for i, tag in enumerate(stream_unit_tags):
            master_sheet.cell(row=3, column=4 + i, value=tag)

        # Map streamtag → row number
        stream_tag_to_row = {
            row[2].value.strip(): row_idx
            for row_idx, row in enumerate(stream_sheet.iter_rows(min_row=3, max_col=3), start=3)
            if row[2].value
        }

        # Get parameters from master sheet: row mapping for each parameter
        param_rows = {}
        syscad_start = None
        for row in range(1, master_sheet.max_row + 1):
            cat_val = master_sheet.cell(row=row, column=1).value
            if cat_val and "SysCAD Inputs" in str(cat_val):
                syscad_start = row + 1
                continue
            if syscad_start and row >= syscad_start:
                val = master_sheet.cell(row=row, column=1).value
                if val and "Engineering Inputs" in str(val):
                    break
                param_name = master_sheet.cell(row=row, column=2).value
                if param_name:
                    param_rows[param_name.strip()] = row

        # Re-fetch master unit tags now that we've populated them
        master_unit_tags = [cell.value for cell in master_sheet[3][3:] if cell.value]

        # Populate values and update units
        for col_offset, unit_tag in enumerate(master_unit_tags):
            if unit_tag not in stream_unit_tags:
                continue
            stream_col = stream_unit_tags.index(unit_tag) + 4  # D=4 in Excel
            for master_param, stream_tag in parameter_mapping.items():
                if master_param not in param_rows or stream_tag not in stream_tag_to_row:
                    continue
                master_row = param_rows[master_param]
                stream_row = stream_tag_to_row[stream_tag]

                value = stream_sheet.cell(row=stream_row, column=stream_col).value
                if value is not None:
                    master_col = col_offset + 4  # D=4
                    master_sheet.cell(row=master_row, column=master_col, value=value)

                # Update unit in master sheet (column C)
                stream_unit = stream_sheet.cell(row=stream_row, column=2).value  # Column B
                master_unit = master_sheet.cell(row=master_row, column=3).value  # Column C
                if stream_unit and master_unit != stream_unit:
                    master_sheet.cell(row=master_row, column=3, value=stream_unit)

    # Save result to BytesIO
    output_stream = BytesIO()
    master_wb.save(output_stream)
    output_stream.seek(0)

    # Optionally return missing types too (could be logged or displayed in Streamlit)
    return output_stream, missing_types
